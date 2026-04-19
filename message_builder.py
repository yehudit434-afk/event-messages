import os
from dataclasses import dataclass
from openpyxl import load_workbook
from dotenv import load_dotenv
from config import EVENTS_FILE, ENV_FILE
from constants import SHEET_TEMPLATES, GREETING_CLOSING, NO_TEMPLATE_ERROR

load_dotenv(ENV_FILE)
FAMILY_SIGNATURE = os.getenv("FAMILY_SIGNATURE")

HTML_TEMPLATE = """\
<html dir="rtl">
<body style="font-family: Arial, sans-serif; text-align: center; direction: rtl; padding: 20px;">
    <div style="max-width: 500px; margin: 0 auto; background: {bg_color}; border-radius: 16px; padding: 30px;">
        <h1 style="font-size: {title_font_size}px; margin-bottom: 20px; white-space: nowrap;">{title}</h1>
        <p style="font-size: 18px; line-height: 1.8; color: #333;">
            {body}
        </p>
        <hr style="border: none; border-top: 1px solid #ccc; margin: 24px 0;">
        <p style="font-size: 16px; color: #555;">
            {closing}<br>
            {signature}
        </p>
    </div>
</body>
</html>
"""


@dataclass
class Event:
    date: str
    name: str
    gender: str
    event_type: str
    baby_name: str = ""


def _load_templates() -> dict:
    """Loads message templates from the 'תבניות' sheet."""
    wb = load_workbook(EVENTS_FILE, read_only=True)
    ws = wb[SHEET_TEMPLATES]
    templates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        event_type = str(row[0])
        gender = str(row[1]) if row[1] else ""
        templates[(event_type, gender)] = {
            "subject": str(row[2]),
            "body": str(row[3]).replace("{br}", "<br>"),
            "bg_color": str(row[4]),
        }
    wb.close()
    return templates


def _find_template(templates: dict, event_type: str, gender: str) -> dict | None:
    """Finds a template by exact match, then by empty gender, then by any matching event type."""
    data = templates.get((event_type, gender))
    if data:
        return data
    data = templates.get((event_type, ""))
    if data:
        return data
    for key, val in templates.items():
        if key[0] == event_type:
            return val
    return None


def build_message(event: Event) -> tuple[str, str]:
    """Builds subject and HTML body for the greeting email."""
    templates = _load_templates()
    data = _find_template(templates, event.event_type, event.gender)
    if not data:
        raise ValueError(NO_TEMPLATE_ERROR.format(event_type=event.event_type, gender=event.gender))

    fmt = {"name": event.name, "baby_name": event.baby_name}
    subject = data["subject"].format(**fmt)
    title_font_size = min(26, int(500 / (len(subject) * 0.6)))
    body = HTML_TEMPLATE.format(
        title=subject,
        title_font_size=title_font_size,
        body=data["body"].format(**fmt),
        bg_color=data["bg_color"],
        closing=GREETING_CLOSING,
        signature=FAMILY_SIGNATURE,
    )
    return subject, body
