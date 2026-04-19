from datetime import date
from openpyxl import load_workbook
from pyluach import dates
from config import EVENTS_FILE
from message_builder import Event
from constants import SHEET_EVENTS, SHEET_TEMP_EVENTS, SHEET_SEND_LOG, SENT_MARKER


def load_recurring_events() -> list[Event]:
    """Loads recurring events from the first sheet."""
    wb = load_workbook(EVENTS_FILE, read_only=True)
    ws = wb[SHEET_EVENTS]
    events = [
        Event(date=str(row[0]), name=str(row[1]), gender=str(row[2]) if row[2] else "",
              event_type=str(row[3]))
        for row in ws.iter_rows(min_row=2, values_only=True) if row[0]
    ]
    wb.close()
    return events


def load_temporary_events() -> list[Event]:
    """Loads unsent temporary events from the second sheet."""
    wb = load_workbook(EVENTS_FILE, read_only=True)
    ws = wb[SHEET_TEMP_EVENTS]
    events = [
        Event(
            date=str(row[0]), name=str(row[1]) if row[1] else "",
            gender=str(row[3]) if row[3] else "", event_type=str(row[4]),
            baby_name=str(row[2]) if row[2] else "",
        )
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] and (not row[5] or str(row[5]) != SENT_MARKER)
    ]
    wb.close()
    return events


def mark_as_sent(event: Event) -> None:
    """Marks a temporary event as sent in the Excel file."""
    wb = load_workbook(EVENTS_FILE)
    ws = wb[SHEET_TEMP_EVENTS]
    for row in ws.iter_rows(min_row=2):
        if (row[0].value and str(row[0].value) == event.date
                and (str(row[1].value) if row[1].value else "") == event.name
                and str(row[4].value) == event.event_type
                and not row[5].value):
            row[5].value = SENT_MARKER
            break
    wb.save(EVENTS_FILE)
    wb.close()


def is_already_sent(event: Event, hebrew_date_str: str) -> bool:
    """Checks if a recurring event was already sent this Hebrew year."""
    current_hebrew_year = dates.HebrewDate.today().year
    wb = load_workbook(EVENTS_FILE, read_only=True)
    ws = wb[SHEET_SEND_LOG]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (row and row[0] and str(row[0]) == hebrew_date_str
                and str(row[1]) == event.name
                and str(row[2]) == event.event_type
                and row[3]):
            sent_date = date.fromisoformat(str(row[3])[:10])
            sent_hebrew_year = dates.HebrewDate.from_pydate(sent_date).year
            if sent_hebrew_year == current_hebrew_year:
                wb.close()
                return True
    wb.close()
    return False


def log_sent(event: Event, hebrew_date_str: str) -> None:
    """Logs a sent recurring event."""
    wb = load_workbook(EVENTS_FILE)
    ws = wb[SHEET_SEND_LOG]
    ws.append([hebrew_date_str, event.name, event.event_type, date.today().isoformat()])
    wb.save(EVENTS_FILE)
    wb.close()
